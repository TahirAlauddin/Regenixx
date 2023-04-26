# IMPORT PACKAGES AND MODULES
# ///////////////////////////////////////////////////////////////
import os
import sqlite3
from generate_pdf import generatePDF
from round_image import add_rounded_border
from threading import Thread
import winreg
from constants import EXCEL_FILE_NAME

# IMPORT QT CORE
# ///////////////////////////////////////////////////////////////
from qt_core import *

# LOAD UI MAIN
# ///////////////////////////////////////////////////////////////
from . ui_main import *
from widgets.py_autocomplete_box import AutocompleteSearchBox

# FUNCTIONS
class MainFunctions():
    def __init__(self):
        super().__init__()
        # SETUP MAIN WINDOw
        # Load widgets from "gui\uis\main_window\ui_main.py"
        # ///////////////////////////////////////////////////////////////
        self.ui = UI_MainWindow()
        self.ui.setup_ui(self)

    # SET MAIN WINDOW PAGES
    # ///////////////////////////////////////////////////////////////
    def set_page(self, page):
        self.ui.load_pages.pages.setCurrentWidget(page)

    # SET LEFT COLUMN PAGES
    # ///////////////////////////////////////////////////////////////
    def set_left_column_menu(
        self,
        menu,
        title,
        icon_path
    ):
        self.ui.left_column.menus.menus.setCurrentWidget(menu)
        self.ui.left_column.title_label.setText(title)
        self.ui.left_column.icon.set_icon(icon_path)

    # RETURN IF LEFT COLUMN IS VISIBLE
    # ///////////////////////////////////////////////////////////////
    def left_column_is_visible(self):
        width = self.ui.left_column_frame.width()
        if width == 0:
            return False
        else:
            return True

    # RETURN IF RIGHT COLUMN IS VISIBLE
    # ///////////////////////////////////////////////////////////////
    def right_column_is_visible(self):
        width = self.ui.right_column_frame.width()
        if width == 0:
            return False
        else:
            return True

    # SET RIGHT COLUMN PAGES
    # ///////////////////////////////////////////////////////////////
    def set_right_column_menu(self, menu):
        self.ui.right_column.menus.setCurrentWidget(menu)

    # GET TITLE BUTTON BY OBJECT NAME
    # ///////////////////////////////////////////////////////////////
    def get_title_bar_btn(self, object_name):
        return self.ui.title_bar_frame.findChild(QPushButton, object_name)

    # GET TITLE BUTTON BY OBJECT NAME
    # ///////////////////////////////////////////////////////////////
    def get_left_menu_btn(self, object_name):
        return self.ui.left_menu.findChild(QPushButton, object_name)
    
    # LEDT AND RIGHT COLUMNS / SHOW / HIDE
    # ///////////////////////////////////////////////////////////////
    def toggle_left_column(self):
        # GET ACTUAL CLUMNS SIZE
        width = self.ui.left_column_frame.width()
        right_column_width = self.ui.right_column_frame.width()

        MainFunctions.start_box_animation(self, width, right_column_width, "left")

    def toggle_right_column(self):
        # GET ACTUAL CLUMNS SIZE
        left_column_width = self.ui.left_column_frame.width()
        width = self.ui.right_column_frame.width()

        MainFunctions.start_box_animation(self, left_column_width, width, "right")

    def start_box_animation(self, left_box_width, right_box_width, direction):
        right_width = 0
        left_width = 0
        time_animation = self.ui.settings["time_animation"]
        minimum_left = self.ui.settings["left_column_size"]["minimum"]
        maximum_left = self.ui.settings["left_column_size"]["maximum"]
        minimum_right = self.ui.settings["right_column_size"]["minimum"]
        maximum_right = self.ui.settings["right_column_size"]["maximum"]

        # Check Left Values        
        if left_box_width == minimum_left and direction == "left":
            left_width = maximum_left
        else:
            left_width = minimum_left

        # Check Right values        
        if right_box_width == minimum_right and direction == "right":
            right_width = maximum_right
        else:
            right_width = minimum_right       

        # ANIMATION LEFT BOX        
        self.left_box = QPropertyAnimation(self.ui.left_column_frame, b"minimumWidth")
        self.left_box.setDuration(time_animation)
        self.left_box.setStartValue(left_box_width)
        self.left_box.setEndValue(left_width)
        self.left_box.setEasingCurve(QEasingCurve.InOutQuart)

        # ANIMATION RIGHT BOX        
        self.right_box = QPropertyAnimation(self.ui.right_column_frame, b"minimumWidth")
        self.right_box.setDuration(time_animation)
        self.right_box.setStartValue(right_box_width)
        self.right_box.setEndValue(right_width)
        self.right_box.setEasingCurve(QEasingCurve.InOutQuart)

        # GROUP ANIMATION
        self.group = QParallelAnimationGroup()
        self.group.stop()
        self.group.addAnimation(self.left_box)
        self.group.addAnimation(self.right_box)
        self.group.start()

               
    def selectImage(self, label):
        file, _ = QFileDialog.getOpenFileName(self, 'Pdf Images', '.', 'Images (*.jpg;*.png;*jpeg)' )
        if file:
            MainFunctions.setImageForPixmap(label, 350, 280, file)

            if label == self.image1LabelPixmap:
                self.image_plan1 = file
            elif label == self.image2LabelPixmap:
                self.image_plan2 = file
            elif label == self.image3LabelPixmap:
                self.image_plan3 = file

        
    # Functions for Add, Remove
    def addRow(tableWidget: QTableWidget):
        rowCount = tableWidget.rowCount()
        tableWidget.setRowCount(rowCount + 1)
        tableWidget.setItem(rowCount + 1, 0, QTableWidgetItem())
        tableWidget.setItem(rowCount + 1, 1, QTableWidgetItem())
        tableWidget.setItem(rowCount + 1, 2, QTableWidgetItem())
        tableWidget.setItem(rowCount + 1, 3, QTableWidgetItem())
        
    def removeRow(tableWidget: QTableWidget):
        tableWidget.removeRow(tableWidget.currentRow())


    def getTableContent(table):
        result = []
        for row in range(table.rowCount()):
            try:
                title = table.item(row, 0).text()
                cost = table.item(row, 1).text()
                quantity = table.item(row, 2).text()
            except:
                # If reading the data from tables throws error, it means the cell were empty
                raise Exception(f'Empty Cell in table {table.objectName()}')
            
            # If either quanity or cost is not a number, throw an error
            if not quantity.isdigit():
                raise Exception(f'Incorrect Data Type for Quantity in table {table.objectName()}.')
            if not cost.isdigit():
                raise Exception(f'Incorrect Data Type for Cost in table {table.objectName()}.')
            
            result.append([title, cost, quantity])
        
        return result
    

    def createPDF(mainWindow):
        mainWindow.pdfCreateSignal.emit()

        # Create PDF
        diagnoses = []
        for diagnosisLineEdit in mainWindow.diagnosesLineEdits:
            diagnoses.append(diagnosisLineEdit.text())
        
        # Write all diagnoses into 1 diagnosis
        diagnosis = ', '.join(diagnoses)
        patients_name = mainWindow.patientsNameLineEdit.text()
        patients_date_of_birth = mainWindow.patientsDateOfBirth.date().toPython().strftime("%B %d, %Y")

        # Get discounts for each plan
        discountPlan1 = int(mainWindow.discountComboBoxBestTable.currentText().strip('%'))
        discountPlan2 = int(mainWindow.discountComboBoxBetterTable.currentText().strip('%'))
        discountPlan3 = int(mainWindow.discountComboBoxGoodTable.currentText().strip('%'))
        
        # Get the data from the tables
        try:
            # Also use list slicing to get the first 8 rows only
            services_plan1 = MainFunctions.getTableContent(mainWindow.bestTable)[:8]
            services_plan2 = MainFunctions.getTableContent(mainWindow.betterTable)[:8]
            services_plan3 = MainFunctions.getTableContent(mainWindow.goodTable)[:8]
        except Exception as e:
            QMessageBox.critical(mainWindow, "Error", f"{str(e)}")
            mainWindow.removeProgressBar.emit()
            return
        
        # Handle wrong/incomplete user input (Text)
        if not all([services_plan1, services_plan2, services_plan3]):
            QMessageBox.critical(mainWindow, "Incomplete Information", f"Please make sure all three tables have atleast 1 row of valid information.")
            mainWindow.removeProgressBar.emit()
            return
        
        if (not patients_name):
            QMessageBox.critical(mainWindow, "Incomplete Information", f"Please enter Patient's Name.")
            mainWindow.removeProgressBar.emit()
            return
        if (not diagnosis):
            QMessageBox.critical(mainWindow, "Incomplete Information", f"Please enter atleast 1 Diagnosis.")
            mainWindow.removeProgressBar.emit()
            return
            

        # Handle wrong/incomplete user input (Images)
        if not all([mainWindow.image_plan1, mainWindow.image_plan2, mainWindow.image_plan3]):
            # User must provide all images
            QMessageBox.critical(mainWindow, "Error occured!", "Please provide all images!")
            mainWindow.removeProgressBar.emit()
            return
            mainWindow.image_plan1, mainWindow.image_plan2, mainWindow.image_plan3 = ['images/consultation2.jpg', 
                                                                                      'images/financialconsultation.jpg',
                                                                                      'images/JENNWIDE2.jpg']
        
        for diagnosis_for_db in diagnoses:
            Database.add_diagnosis_in_db(diagnosis_for_db)

        diagnoses = Database.get_diagnosis_from_db()
        mainWindow.SetupMainWindow.removeCustomWidgetFromWindow(mainWindow)
        mainWindow.diagnosisLineEdit = AutocompleteSearchBox(
            text = "",
            place_holder_text = "Diagnosis Description",
            radius = 8,
            border_size = 2,
            color = mainWindow.themes["app_color"]["text_foreground"],
            selection_color = mainWindow.themes["app_color"]["white"],
            bg_color = mainWindow.themes["app_color"]["dark_one"],
            bg_color_active = mainWindow.themes["app_color"]["dark_three"],
            context_color = mainWindow.themes["app_color"]["context_color"],
            items=diagnoses
        )
        mainWindow.SetupMainWindow.addWidgetToWindow(mainWindow)

        def populateDataIntoPDF():
            if not os.path.exists('pdf-images'):
                os.mkdir('pdf-images')
            mainWindow.pdfCreateShowProgressSignal.emit('CROPPING IMAGE 1')
            image_plan1 = add_rounded_border(mainWindow.image_plan1)
            mainWindow.pdfCreateShowProgressSignal.emit('CROPPING IMAGE 2')
            image_plan2 = add_rounded_border(mainWindow.image_plan2)
            mainWindow.pdfCreateShowProgressSignal.emit('CROPPING IMAGE 3')
            image_plan3 = add_rounded_border(mainWindow.image_plan3)
            mainWindow.pdfCreateShowProgressSignal.emit('IMAGES CROPPED SUCCESSFULLY!')
            # Generate PDF File
            try:
                output_folder = MainFunctions.get_users_desktop_folder()
                pdf_path = os.path.join(output_folder, f'{patients_name} - {patients_date_of_birth}.pdf')
                generatePDF(patients_name, patients_date_of_birth,
                            diagnosis, image_plan1, image_plan2, image_plan3,
                            services_plan1, services_plan2, services_plan3,
                            discountPlan1, discountPlan2, discountPlan3,
                            output_pdf_path=pdf_path,
                            signal=mainWindow.pdfCreateShowProgressSignal)
                
                # Remove temporary images
                os.remove(image_plan1)
                os.remove(image_plan2)
                os.remove(image_plan3)

                # Add Patient into database
                Database.add_patient_to_database(patients_name, patients_date_of_birth,
                            diagnosis, services_plan1, services_plan2, services_plan3,
                            discountPlan1, discountPlan2, discountPlan3)
                
                mainWindow.patientAddedToDatabaseSignal.emit()

            except Exception as e:
                MainFunctions.log_db_errors(str(e), 'error.log')
            
        thread = Thread(target=populateDataIntoPDF)
        thread.start()


    def get_users_desktop_folder():
        # Open the registry key for the desktop folder
        key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, r"Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders")

        # Get the value of the "Desktop" key
        desktop_path = winreg.QueryValueEx(key, "Desktop")[0]

        return desktop_path

    
    def get_services_from_excel(window):
        import openpyxl
        workbook = openpyxl.load_workbook(EXCEL_FILE_NAME)
        sheet = workbook.active
        
        # Read data from each row
        data = []
        header = []
        for row in sheet.iter_rows(max_row=1, values_only=True):
            try:
                header.extend([
                    str(row[0]),
                    str(row[1]),
                    str(row[2]),
                    str(row[3]),
                ])
            except IndexError:
                header.extend(['Category', 'Title', 'Cost', 'Description'])

        error_occured = False
        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                row_data = [
                    str(row[0]),
                    str(row[1]),
                    str(row[2]),
                    str(row[3]),
                ]
                if str(row[2]).isdigit():
                    data.append(row_data)
            except IndexError:
                if len(data) < 1:
                    data.append(['', '', '', ''])
                    error_occured = True

        if error_occured:
            QMessageBox.critical(window, 'Error in Excel File', 
                            'The program couldn\'t read the Excel file. Make sure there are exactly 4 columns.')

        return data, header

    # Set Image and define the width/height for Pixmap        
    def setImageForPixmap(label, width, height, file=None):
        pixmap = QPixmap(file)
        pixmap = pixmap.scaledToWidth(width , Qt.SmoothTransformation)
        pixmap = pixmap.scaledToHeight(height, Qt.SmoothTransformation)
        label.setPixmap(pixmap)
        label.setMaximumSize(QSize(width, height))

    def changeExcelFile(window):
        filename, _ = QFileDialog.getOpenFileName(window, 'Select Excel File', '', 'Excel Files (*.xlsx)')
        if filename:
            filecontent = open(filename, 'rb').read()
            open(EXCEL_FILE_NAME, 'wb').write(filecontent)

    def log_db_errors(message=None, log_file='db.error.log'):
        import traceback, datetime
        # Get error message with traceback
        error_message = traceback.format_exc()
        # Format error message with timestamp
        log_message = f"{datetime.datetime.now()} - ERROR - {error_message}"
        with open(log_file, 'a') as error_log_file:
            error_log_file.write(log_message)
            if message:
                error_log_file.write(message + '\n')
                



import sqlite3
from datetime import datetime


class Database:

    def add_patient_to_database(patient_name, date_of_birth, diagnosis,
                                services_plan1, services_plan2, services_plan3,
                                discountPlan1, discountPlan2, discountPlan3):
        try:

            # get current date and time
            current_date = datetime.now()

            # format current date as string in YYYY-MM-DD format
            date_added = current_date.strftime('%Y-%m-%d')

            connection = sqlite3.connect('db.sqlite3')
            cursor = connection.cursor()
            
            # Start transaction
            cursor.execute('BEGIN')

            services_plan1_id = Database.add_service_plan(cursor, services_plan1, discountPlan1)
            services_plan2_id = Database.add_service_plan(cursor, services_plan2, discountPlan2)
            services_plan3_id = Database.add_service_plan(cursor, services_plan3, discountPlan3)

            cursor.execute("CREATE TABLE IF NOT EXISTS Patients (id INTEGER PRIMARY KEY, patients_name VARCHAR, \
                        patients_date_of_birth DATE, date_added DATE, diagnosis VARCHAR, \
                        services_plan1_id INTEGER REFERENCES ServicesPlan(id), \
                        services_plan2_id INTEGER REFERENCES ServicesPlan(id), \
                        services_plan3_id INTEGER REFERENCES ServicesPlan(id))")
            if diagnosis:
                cursor.execute("INSERT INTO Patients (patients_name, patients_date_of_birth, date_added,\
                                diagnosis, services_plan1_id, services_plan2_id, services_plan3_id) \
                                VALUES (?, ?, ?, ?, ?, ?, ?)", 
                (patient_name, date_of_birth, date_added, diagnosis, services_plan1_id, services_plan2_id, services_plan3_id))


            # End transaction
            connection.commit()

        except Exception as e:
            # Rollback transaction in case of errors
            cursor.execute("ROLLBACK")
            MainFunctions.log_db_errors(str(e))
            return False
        finally:
            connection.close()
        return True
    

    def add_service_plan(cursor, services, discount):
        
        cursor.execute("CREATE TABLE IF NOT EXISTS ServicesPlan (id INTEGER PRIMARY KEY, discount INTEGER)")
        cursor.execute("CREATE TABLE IF NOT EXISTS Services (id INTEGER PRIMARY KEY, service_plan_id INTEGER REFERENCES ServicesPlan(id), service VARCHAR, quantity INEGER, price INTEGER)")

        # Insert service plan into ServicesPlan table
        cursor.execute('INSERT INTO ServicesPlan (discount) VALUES (?)', (discount,)) 
        service_plan_id = cursor.lastrowid

        # Insert each service into Services table
        for service, quantity, price in services:
            cursor.execute('INSERT INTO Services (service_plan_id, service, quantity, price) VALUES (?, ?, ?, ?)',
                            (service_plan_id, service, quantity, price))

        return service_plan_id
    

    def get_services():
        try:
            connection = sqlite3.connect('db.sqlite3')
            cursor = connection.cursor()

            # Fetch all service plans with their associated services
            cursor.execute('SELECT ServicesPlan.id, Services.service, Services.quantity, Services.price, ServicesPlan.discount \
                            FROM ServicesPlan JOIN Services ON Services.service_plan_id = ServicesPlan.id')
            service_data = cursor.fetchall()

            # Transform data into a dictionary where each service plan is a key and its value is a list of associated services
            services = {}
            for row in service_data:
                service_plan_id, service, quantity, price, discount = row
                if service_plan_id not in services:
                    services[service_plan_id] = {'discount': discount, 'services': []}
                services[service_plan_id]['services'].append({'service': service, 'quantity': quantity, 'price': price})


        except Exception as e:
            MainFunctions.log_db_errors(str(e))
            services = []
        finally:
            cursor.close()
            connection.close()
        return services


    def get_patients():
        try:
            connection = sqlite3.connect('db.sqlite3')
            cursor = connection.cursor()

            cursor.execute("CREATE TABLE IF NOT EXISTS ServicesPlan (id INTEGER PRIMARY KEY, discount INTEGER)")
            cursor.execute("CREATE TABLE IF NOT EXISTS Services (id INTEGER PRIMARY KEY, service_plan_id INTEGER REFERENCES ServicesPlan(id), service VARCHAR, quantity INTEGER, price INTEGER)")

            cursor.execute("CREATE TABLE IF NOT EXISTS Patients (id INTEGER PRIMARY KEY, patients_name VARCHAR, \
                        patients_date_of_birth DATE, date_added DATE, diagnosis VARCHAR, \
                        services_plan1_id INTEGER REFERENCES ServicesPlan(id), \
                        services_plan2_id INTEGER REFERENCES ServicesPlan(id), \
                        services_plan3_id INTEGER REFERENCES ServicesPlan(id))")

            # Fetch all patients with their associated service plans
            cursor.execute('SELECT Patients.id, Patients.patients_name, Patients.patients_date_of_birth, \
                            Patients.date_added, Patients.diagnosis, \
                            ServicesPlan1.id, ServicesPlan1.discount, \
                            ServicesPlan2.id, ServicesPlan2.discount, \
                            ServicesPlan3.id, ServicesPlan3.discount \
                            FROM Patients \
                            JOIN ServicesPlan AS ServicesPlan1 ON Patients.services_plan1_id = ServicesPlan1.id \
                            JOIN ServicesPlan AS ServicesPlan2 ON Patients.services_plan2_id = ServicesPlan2.id \
                            JOIN ServicesPlan AS ServicesPlan3 ON Patients.services_plan3_id = ServicesPlan3.id')
            patient_data = cursor.fetchall()

            # Transform data into a list of dictionaries where each dictionary represents a patient
            patients = []
            for row in patient_data:
                patient_id, patient_name, date_of_birth, date_added, diagnosis, \
                service_plan1_id, discount1, service_plan2_id, discount2, service_plan3_id, discount3 = row

                patient = {'id': patient_id,
                        'name': patient_name,
                        'date_of_birth': date_of_birth,
                        'date_added': date_added,
                        'diagnosis': diagnosis,
                        'service_plans': [
                            {'services': Database.get_services_by_service_plan_id(service_plan1_id), 'discount': discount1},
                            {'services': Database.get_services_by_service_plan_id(service_plan2_id), 'discount': discount2},
                            {'services': Database.get_services_by_service_plan_id(service_plan3_id), 'discount': discount3}
                        ]
                        }

                patients.append(patient)


        except Exception as e:
            MainFunctions.log_db_errors(str(e))
            patients = []
        finally:
            cursor.close()
            connection.close()
        return patients


    def delete_patient(patient_id):
        try:
            connection = sqlite3.connect('db.sqlite3')
            cursor = connection.cursor()
            result = cursor.execute("SELECT services_plan1_id, services_plan2_id, services_plan3_id FROM Patients WHERE id = ?", 
                                    (patient_id,))
            services_plan1_id, services_plan2_id, services_plan3_id = result.fetchall()[0]
            cursor.execute("DELETE FROM Patients WHERE id = ?", (patient_id,))
            cursor.execute("DELETE FROM ServicesPlan WHERE id = ?", (services_plan1_id,))
            cursor.execute("DELETE FROM ServicesPlan WHERE id = ?", (services_plan2_id,))
            cursor.execute("DELETE FROM ServicesPlan WHERE id = ?", (services_plan3_id,))
            connection.commit()

        except Exception as e:
            MainFunctions.log_db_errors(str(e))

        finally:
            cursor.close()
            connection.close()


    def get_services_by_service_plan_id(service_plan_id):
        try:
            connection = sqlite3.connect('db.sqlite3')
            cursor = connection.cursor()

            # Select all services for the given service plan ID
            cursor.execute('SELECT service, quantity, price FROM Services WHERE service_plan_id = ?', (service_plan_id,))
            services = cursor.fetchall()

        except Exception as e:
            MainFunctions.log_db_errors(str(e))
            services = []
        finally:
            cursor.close()
            connection.close()
        return services


    def get_diagnosis_from_db():
        try:
            connection = sqlite3.connect('db.sqlite3')
            cursor = connection.cursor()
            cursor.execute("CREATE TABLE IF NOT EXISTS Diagnosis (id INTEGER PRIMARY KEY, diagnosis VARCHAR UNIQUE)")
            cursor.execute("SELECT diagnosis FROM Diagnosis")
            diagnoses = [item[0] for item in cursor.fetchall()]
        except Exception as e:
            MainFunctions.log_db_errors(str(e))
            diagnoses = []
        finally:
            # Close cursor and connection
            cursor.close()
            connection.close()
        return diagnoses


    def add_diagnosis_in_db(diagnosis: str):
        added = False
        try:
            connection = sqlite3.connect('db.sqlite3')
            cursor = connection.cursor()
            cursor.execute("CREATE TABLE IF NOT EXISTS Diagnosis (id INTEGER PRIMARY KEY, diagnosis VARCHAR UNIQUE)")
            if diagnosis:
                cursor.execute("INSERT INTO Diagnosis (diagnosis) VALUES (?)", (diagnosis,))
            added = True
            connection.commit()
        except Exception as e:
            MainFunctions.log_db_errors(str(e))
        finally:
            # Close cursor and connection
            cursor.close()
            connection.close()
        return added


    def update_diagnoses_in_db(diagnoses):
        successfully_updated = False
        conn = sqlite3.connect('db.sqlite3')
        cursor = conn.cursor()
        # Begin transaction
        try:
            cursor.execute("BEGIN")

            cursor.execute("DELETE FROM Diagnosis;")

            # Insert data into Diagnosis
            cursor.executemany("INSERT INTO Diagnosis (id, diagnosis) VALUES (?, ?)", diagnoses)

            # Commit transaction if all statements are successful
            conn.commit()
            print("Transaction completed successfully.")
            successfully_updated = True

        except:
            # Rollback transaction if any statement fails
            conn.rollback()
            MainFunctions.log_db_errors("Transaction failed. Rolled back changes.")
            successfully_updated = False

        finally:
            # Close cursor and connection
            cursor.close()
            conn.close()

        return successfully_updated


    def list_diagnosis_from_db():
        try:
            connection = sqlite3.connect('db.sqlite3')
            cursor = connection.cursor()
            cursor.execute("CREATE TABLE IF NOT EXISTS Diagnosis (id INTEGER PRIMARY KEY, diagnosis VARCHAR UNIQUE)")
            cursor.execute("SELECT id, diagnosis FROM Diagnosis")
            diagnoses = [(item[0], item[1]) for item in cursor.fetchall()]
            return diagnoses
        except:
            MainFunctions.log_db_errors()
        finally:
            # Close cursor and connection
            cursor.close()
            connection.close()
            